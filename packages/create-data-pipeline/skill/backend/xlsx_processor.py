"""
Excel XLSX 文件处理器 - 提取 Excel 表格数据
"""

import logging
from pathlib import Path
from typing import List, Dict, Any
from .file_processor import ProcessedDocument

logger = logging.getLogger(__name__)


class XlsxProcessor:
    """Excel XLSX 工作表处理器"""

    def __init__(self):
        self.clean_empty_rows = True  # 是否清理空行
        self.clean_empty_cols = True  # 是否清理空列

    def process(self, file_path: str) -> ProcessedDocument:
        """处理 Excel XLSX 文件"""
        try:
            import openpyxl
        except ImportError:
            return ProcessedDocument(
                content="",
                errors=["缺少依赖: openpyxl。请运行: pip install openpyxl"]
            )

        path = Path(file_path)
        errors = []
        content_parts = []
        tables = []

        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)

            # 提取元数据
            metadata = {
                "source": str(path),
                "type": "xlsx",
                "sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
            }

            # 处理每个工作表
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    table_content = self._process_worksheet(ws, sheet_name)

                    if table_content.strip():
                        content_parts.append(f"\\n## 工作表: {sheet_name}\\n")
                        content_parts.append(table_content)

                        # 同时保存为独立的表格
                        tables.append(table_content)

                except Exception as e:
                    errors.append(f"工作表 {sheet_name} 处理失败: {str(e)}")

            content = "\\n".join(content_parts)

            return ProcessedDocument(
                content=content,
                metadata=metadata,
                tables=tables,
                errors=errors
            )

        except Exception as e:
            logger.error(f"XLSX 处理失败: {e}")
            return ProcessedDocument(
                content="",
                errors=[f"XLSX 处理失败: {str(e)}"]
            )

    def _process_worksheet(self, ws, sheet_name: str) -> str:
        """处理单个工作表"""
        try:
            if self.clean_empty_rows:
                rows = self._clean_empty_rows(ws)
            else:
                rows = list(ws.rows)

            if not rows:
                return ""

            # 转换为 Markdown 表格
            return self._rows_to_markdown(rows, sheet_name)

        except Exception as e:
            logger.error(f"工作表处理失败: {e}")
            return ""

    def _clean_empty_rows(self, ws) -> List:
        """清理完全空的行"""
        cleaned_rows = []
        for row in ws.rows:
            # 检查行是否有任何数据
            has_data = any(cell.value is not None and str(cell.value).strip()
                          for cell in row)
            if has_data:
                if self.clean_empty_cols:
                    # 同时清理空列
                    cleaned_row = [cell for cell in row
                                  if cell.value is not None and str(cell.value).strip()]
                    if cleaned_row:
                        cleaned_rows.append(cleaned_row)
                else:
                    cleaned_rows.append(row)

        return cleaned_rows

    def _rows_to_markdown(self, rows: List, sheet_name: str) -> str:
        """将行数据转换为 Markdown 表格"""
        if not rows:
            return ""

        markdown_lines = []

        for row in rows:
            row_data = []
            for cell in row:
                # 获取单元格值
                value = cell.value if cell.value is not None else ""
                # 转换为字符串
                text = str(value).strip()
                row_data.append(text)

            if row_data:  # 只有行有数据时才添加
                markdown_lines.append("| " + " | ".join(row_data) + " |")

        if markdown_lines:
            # 添加表头分隔线
            max_cols = max(len(row) if row else 0 for row in rows)
            markdown_lines.insert(1, "|" + "|".join(["---"] * max_cols) + "|")

        return "\\n".join(markdown_lines)
"""
Excel XLSX 文件处理器 - 提取 Excel 表格数据
"""

import logging
from pathlib import Path
from typing import List, Dict, Any
from .file_processor import ProcessedDocument

logger = logging.getLogger(__name__)


class XlsxProcessor:
    """Excel XLSX 工作表处理器"""

    def __init__(self):
        self.clean_empty_rows = True  # 是否清理空行
        self.clean_empty_cols = True  # 是否清理空列

    def process(self, file_path: str) -> ProcessedDocument:
        """处理 Excel XLSX 文件"""
        try:
            import openpyxl
        except ImportError:
            return ProcessedDocument(
                content="",
                errors=["缺少依赖: openpyxl。请运行: pip install openpyxl"]
            )

        path = Path(file_path)
        errors = []
        content_parts = []
        tables = []

        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)

            # 提取元数据
            metadata = {
                "source": str(path),
                "type": "xlsx",
                "sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
            }

            # 处理每个工作表
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    table_content = self._process_worksheet(ws, sheet_name)

                    if table_content.strip():
                        content_parts.append(f"\\n## 工作表: {sheet_name}\\n")
                        content_parts.append(table_content)

                        # 同时保存为独立的表格
                        tables.append(table_content)

                except Exception as e:
                    errors.append(f"工作表 {sheet_name} 处理失败: {str(e)}")

            content = "\\n".join(content_parts)

            return ProcessedDocument(
                content=content,
                metadata=metadata,
                tables=tables,
                errors=errors
            )

        except Exception as e:
            logger.error(f"XLSX 处理失败: {e}")
            return ProcessedDocument(
                content="",
                errors=[f"XLSX 处理失败: {str(e)}"]
            )

    def _process_worksheet(self, ws, sheet_name: str) -> str:
        """处理单个工作表"""
        try:
            if self.clean_empty_rows:
                rows = self._clean_empty_rows(ws)
            else:
                rows = list(ws.rows)

            if not rows:
                return ""

            # 转换为 Markdown 表格
            return self._rows_to_markdown(rows, sheet_name)

        except Exception as e:
            logger.error(f"工作表处理失败: {e}")
            return ""

    def _clean_empty_rows(self, ws) -> List:
        """清理完全空的行"""
        cleaned_rows = []
        for row in ws.rows:
            # 检查行是否有任何数据
            has_data = any(cell.value is not None and str(cell.value).strip()
                          for cell in row)
            if has_data:
                if self.clean_empty_cols:
                    # 同时清理空列
                    cleaned_row = [cell for cell in row
                                  if cell.value is not None and str(cell.value).strip()]
                    if cleaned_row:
                        cleaned_rows.append(cleaned_row)
                else:
                    cleaned_rows.append(row)

        return cleaned_rows

    def _rows_to_markdown(self, rows: List, sheet_name: str) -> str:
        """将行数据转换为 Markdown 表格"""
        if not rows:
            return ""

        markdown_lines = []

        for row in rows:
            row_data = []
            for cell in row:
                # 获取单元格值
                value = cell.value if cell.value is not None else ""
                # 转换为字符串
                text = str(value).strip()
                row_data.append(text)

            if row_data:  # 只有行有数据时才添加
                markdown_lines.append("| " + " | ".join(row_data) + " |")

        if markdown_lines:
            # 添加表头分隔线
            max_cols = max(len(row) if row else 0 for row in rows)
            markdown_lines.insert(1, "|" + "|".join(["---"] * max_cols) + "|")

        return "\\n".join(markdown_lines)
